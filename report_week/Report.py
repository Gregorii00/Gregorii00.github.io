import pandas as pd
import openpyxl
import datetime
import os.path
from openpyxl.writer.excel import save_workbook
def week_report(FILE_NAME1, FILE_NAME2, FILE_NAME3, start, end):
    Path_full = 'C:/Users/zhevn/Downloads/'
    FILE_NAME1 = Path_full + FILE_NAME1
    FILE_NAME2 = Path_full + FILE_NAME2
    FILE_NAME3 = Path_full + FILE_NAME3
    training = ['Обучение 1', 'Обучение 2', 'Обучение 3', 'Обучение ICL']
    if os.path.exists(FILE_NAME1) != True:
        return 0
    df = pd.read_excel(FILE_NAME1, sheet_name='Лист')
    df2 = pd.read_excel(FILE_NAME1, sheet_name='Лист2')
    df_aht = pd.read_excel(FILE_NAME2, sheet_name='Лист')
    df_missed = pd.read_excel(FILE_NAME3, sheet_name='Лист')
    wb = openpyxl.load_workbook(FILE_NAME1)
    df = df[['Время', 'Оператор', 'Оценка']]
    df2 = df2[['Время', 'Оценка', 'Оператор', 'Итог']]
    df_aht = df_aht[['Оператор', 'AHT вх., сек']]
    df_missed = df_missed[['Пользователь', 'Пропущено']]
    start = start.replace("-", "")
    end = end.replace("-", "")
    temp = pd.DatetimeIndex(df['Время'])
    df['Date'] = temp.date
    df['Time'] = temp.time
    del df['Время']
    start = datetime.datetime.strptime(start, "%Y%m%d").date()
    end = datetime.datetime.strptime(end, "%Y%m%d").date()
    dates = str(start.day)+'.'+str(start.month) + '-' + str(end.day)+'.'+str(end.month)
    df = df[(df['Date'] >= start) & (df['Date'] <= end) & (pd.isnull(df['Оценка'])!=True)]
    list1 = []
    for under_tems, tems in df.items():
        list1.append(tems)
    j=0
    operators={}
    operators_sum={}
    operators_args = {}
    teacher= []
    summ=0
    for i in df.index:
        operators_sum[list1[0][i]] = operators_sum.get(list1[0][i], 0) + int(list1[1][i])
        operators[list1[0][i]] = operators.get(list1[0][i], 0) + 1
        summ+=1
        if list1[0][i] in training:
            a = []
            a.append(list1[0][i])
            a.append(list1[1][i])
            a.append(list1[2][i])
            a.append(list1[3][i])
            teacher.append(a)
    operators2= operators.copy()
    operators_sum2= operators_sum.copy()
    for i in operators_sum:
        operators_args[i] = operators_sum.get(i, 0)
    for i in operators:
        operators_args[i] = operators_args.get(i, 0)/operators.get(i, 0)
    list2 = []
    for under_tems, tems in df2.items():
        list2.append(tems)
    list3 = []
    for under_tems, tems in df_aht.items():
        list3.append(tems)
    operators_aht = {}
    for i in df_aht.index:
       operators_aht[list3[0][i]] = list3[1][i]
    list4 = []
    for under_tems, tems in df_missed.items():
        list4.append(tems)
    operators_missed = {}
    for i in df_missed.index:
       operators_missed[list4[0][i]] = list4[1][i]
    for i in range(0, len(list2[3])):
        if list2[3][i] == 'Необоснованная':
            operators_sum[list2[2][i]] = operators_sum.get(list2[2][i], 0)-list2[1][i]
            operators[list2[2][i]] = operators.get(list2[2][i], 0)-1
    ws = wb.create_sheet(dates)
    ws.cell(row=1, column=1).value = 'Оператор'
    ws.cell(row=1, column=2).value = 'Оценка сумм до перерасчета'
    ws.cell(row=1, column=3).value = 'Количество до перерасчета'
    ws.cell(row=1, column=4).value = 'Оценка до перерасчета'
    ws.cell(row=1, column=5).value = 'Если нужен перерасчет'
    ws.cell(row=1, column=6).value = 'Сумма всех оценок'
    ws.cell(row=1, column=7).value = 'колличество оценок'
    ws.cell(row=1, column=8).value = 'Оценка'
    ws.cell(row=1, column=9).value = 'Если нужен перерасчет'
    ws.cell(row=1, column=10).value = 'AHT'
    ws.cell(row=1, column=11).value = 'Пропущенные'
    row = 2
    for i in operators:
        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = operators_sum2.get(i, 0)
        ws.cell(row=row, column=3).value = operators2.get(i, 0)
        ws.cell(row=row, column=4).value = operators_args.get(i, 0)
        ws.cell(row=row, column=6).value = operators_sum.get(i, 0)
        ws.cell(row=row, column=7).value = operators.get(i, 0)
        ws.cell(row=row, column=8).value = operators_sum.get(i, 0)/float(operators.get(i, 0))
        ws.cell(row=row, column=10).value = operators_aht.get(i, 0)
        ws.cell(row=row, column=11).value = operators_missed.get(i, 0)
        row += 1
    teacher2=[]
    if teacher != []:
        for i in training:
            ws = wb.create_sheet(dates + i)
            ws.cell(row=1, column=1).value = 'Дата'
            ws.cell(row=1, column=2).value = 'Час'
            ws.cell(row=1, column=3).value = 'Сумма оценок'
            ws.cell(row=1, column=4).value = 'Количество'
            row = 2
            for j in teacher:
                if j[0]==i:
                    ws.cell(row=row, column=1).value = j[2]
                    ws.cell(row=row, column=2).value = j[3]
                    ws.cell(row=row, column=3).value = j[1]
                    ws.cell(row=row, column=4).value = 1
                    row += 1
                    
    save_workbook(wb, FILE_NAME1)
    return summ
